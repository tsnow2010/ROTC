import landnavTrainer
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
import os
import pandas as pd

## Run program ##

# Create a directory for assignment documents and change directory.
os.makedirs("Assignments", exist_ok=True)
os.chdir(f'{landnavTrainer.home_path}/Assignments')

# Ask for input from user.
train_name = input('What is the name of this training event? (e.g. FTX) ')
num_groups = int(input('How many groups do you have? (e.g. 2) '))
for i in range(1,num_groups+1):
    group = input(f"What is the name of group {i}?  (e.g. MS1&2, MS3, etc) ")
    num_cdts = int(input("How many cadets?  (e.g. 15) "))
    num_train_points = int(input("How many points will they be assigned?  (e.g. 8) "))
    landnavTrainer.generate_assignments(group, num_cdts, num_train_points)


# Create a Pandas Excel writer and combine all .csv files into (1) Excel workbook with proper formating.
output_excel = f'{train_name}_Assignments_Compiled.xlsx'

with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
    for filename in os.listdir(os.getcwd()):
        if filename.endswith('.csv'):
            file_path = os.path.join(os.getcwd(), filename)
            sheet_name = os.path.splitext(filename)[0][:31]  # Excel sheet names max out at 31 characters

            # Read CSV and write to Excel sheet
            df = pd.read_csv(file_path)
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

            ws = writer.sheets[sheet_name]

            # --- Add custom header in first row ---
            header_text = sheet_name
            ws.merge_cells("A1:B1")
            cell = ws["A1"]
            cell.value = header_text

            # --- Change cell width ---
            for col_idx, col in enumerate(df.columns, 2):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 13 # all but first column
            ws.column_dimensions['A'].width = 7 # first column

            # --- Change cell height ---
            for i in range(1,len(df)+2):
                ws.row_dimensions[i].height = 35 # all rows

            # --- Change cell formatting ---
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for row in ws.iter_rows(min_row=1, max_row=len(df)+2, min_col=1, max_col=len(df.columns)):
                for cell in row:
                    # --- Change font size ---
                    cell.font = Font(name="Calibri", size=10)
                    # --- Add borders ---
                    cell.border = border

print(f"All assignments and answer keys have been combined into {output_excel}")
