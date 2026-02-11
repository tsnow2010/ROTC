# This program creates one MS Excel workbook that can easily be used to provide land navigation point assignments for ROTC training.
# Requires properly formated CSV files.  
# Please see the repository for examples.

# By Tyler Snow

import itertools
import random
import csv
import json
import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
import os

# Set paths for changing directories
home_path = os.getcwd()
assignments_path = f"{home_path}/'Assignments'"


# Generates assignment list and corresponding answer key in .csv format.
def generate_assignments(group:str, num_cdts:int, num_train_points:int):
    
    # Constants
    group = group # Group name
    num_cdts = num_cdts # Number of cadets in group
    num_train_points = num_train_points # Number of points trainees are looking for
    
    # Step 1: Pull land navigation points from CSV file and create dictionary.
    points = {}
    with open(f'{home_path}/LN_Points.csv', 'r') as file:
        reader = csv.reader(file)
        next(reader)
        for row in reader:
            points[row[2]] = [{'key': row[1]}, {'coordinates': row[0]}]
            
    tot_num_points = len(points) # total number of points in CSV file
    
    # Step 2: Randomize the sets of points that each cadet is given
    # Create unique combinations of points into sets
    combinations = itertools.combinations(range(1,tot_num_points),int(num_train_points))
    
    # Sets of points that will be distributed to cadets
    point_decks = random.sample(sorted(combinations),num_cdts)
    
    # Step 3: Create CSV file with cadet point assignments
    with open(f'{group} Assignments.csv','w', newline= '') as file:
        writer = csv.writer(file)
        for i in range(0,num_cdts):
            point_deck = point_decks[i]
            coordinates = []
            for point in point_deck:
                coordinates.append(points[str(point)][1]['coordinates'])
            coordinates.insert(0, f'Cadet {i+1}')
            writer.writerow(coordinates)
            writer.writerow(['Answer']) # Add line for writing answers   + ','*(num_train_points-1)]
    
    # Step 4: Create CSV file with cadet point answer key
    with open(f'{group} (Answer Key).csv','w', newline= '') as file:
        writer = csv.writer(file)
        for i in range(0,num_cdts):
            point_deck = point_decks[i]
            keys = []
            for point in point_deck:
                keys.append(points[str(point)][0]['key'] + f' ({point})')
            keys.insert(0, f'Cadet {i+1}')
            writer.writerow(keys)
            
## Run program ##

# Create a directory for assignment documents and change directory.
os.makedirs("Assignments", exist_ok=True)
os.chdir(f'{home_path}/Assignments')

# Ask for input from user.
train_name = input('What is the name of this training event? (e.g. FTX) ')
num_groups = int(input('How many groups do you have? (e.g. 2) '))
for i in range(1,num_groups+1):
    group = input(f"What is the name of group {i}?  (e.g. MS1&2, MS3, etc) ")
    num_cdts = int(input("How many cadets?  (e.g. 15) "))
    num_train_points = int(input("How many points will they be assigned?  (e.g. 8) "))
    generate_assignments(group, num_cdts, num_train_points)


# Create a Pandas Excel writer and combine all .csv files into (1) Excel workbook with proper formating.
output_excel = f'{train_name}_Assignments_Compiled.xlsx'

with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
    for filename in os.listdir(os.getcwd()):
        if filename.endswith('.csv'):
            file_path = os.path.join(os.getcwd(), filename)
            sheet_name = os.path.splitext(filename)[0][:31]  # Excel sheet names max out at 31 characters

            # Read CSV and write to Excel sheet
            df = pd.read_csv(file_path)
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

            ws = writer.sheets[sheet_name]

            # --- Add custom header in first row ---
            header_text = sheet_name
            ws.merge_cells("A1:B1")
            cell = ws["A1"]
            cell.value = header_text

            # --- Change cell width ---
            for col_idx, col in enumerate(df.columns, 2):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 13 # all but first column
            ws.column_dimensions['A'].width = 7 # first column

            # --- Change cell height ---
            for i in range(1,len(df)+2):
                ws.row_dimensions[i].height = 35 # all rows

            # --- Change cell formatting ---
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        
            for row in ws.iter_rows(min_row=1, max_row=len(df)+2, min_col=1, max_col=len(df.columns)):
                for cell in row:
                    # --- Change font size ---
                    cell.font = Font(name="Calibri", size=10)
                    # --- Add borders ---
                    cell.border = border

print(f"All assignments and answer keys have been combined into {output_excel}")
    
    